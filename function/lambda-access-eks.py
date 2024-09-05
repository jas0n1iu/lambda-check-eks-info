import os
import json
import boto3
import base64
import re
import io
import openpyxl
import datetime
from botocore.signers import RequestSigner
from kubernetes import client, config
from kubernetes.client.rest import ApiException
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

STS_TOKEN_EXPIRES_IN = 60
session = boto3.session.Session()
sts = session.client('sts')
service_id = sts.meta.service_model.service_id
regions = os.environ.get('REGIONS').split(',')
lambda_role_arn = os.environ.get('LAMBDA_ROLE_ARN')

def get_bearer_token(cluster_name, region):
    """Create authentication token"""
    eks_client = boto3.client('eks', region_name=region)
    signer = RequestSigner(
        service_id,
        region,
        'sts',
        'v4',
        session.get_credentials(),
        session.events
    )
    params = {
        'method': 'GET',
        'url': 'https://sts.{}.amazonaws.com/'
               '?Action=GetCallerIdentity&Version=2011-06-15'.format(region),
        'body': {},
        'headers': {
            'x-k8s-aws-id': cluster_name
        },
        'context': {}
    }

    signed_url = signer.generate_presigned_url(
        params,
        region_name=region,
        expires_in=STS_TOKEN_EXPIRES_IN,
        operation_name=''
    )
    base64_url = base64.urlsafe_b64encode(signed_url.encode('utf-8')).decode('utf-8')

    # remove any base64 encoding padding:
    return 'k8s-aws-v1.' + re.sub(r'=*', '', base64_url)

def get_cluster_conf(eks_client, cluster_name, region):
    """Retrieve cluster endpoint and certificate"""
    cluster_info = eks_client.describe_cluster(name=cluster_name)
    kubeconfig = {
        'apiVersion': 'v1',
        'clusters': [{
            'name': 'cluster1',
            'cluster': {
                'certificate-authority-data': cluster_info['cluster']['certificateAuthority']['data'],
                'server': cluster_info['cluster']['endpoint']}
        }],
        'contexts': [{'name': 'context1', 'context': {'cluster': 'cluster1', 'user': 'lambda'}}],
        'current-context': 'context1',
        'kind': 'Config',
        'preferences': {},
        'users': [{'name': 'lambda', 'user': {'token': get_bearer_token(cluster_name, region)}}]
    }

    return kubeconfig

def lambda_handler(event, context):
    bucket_name = os.environ.get('S3_BUCKET_NAME')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.sheet_view.showGridLines = False  # Hide gridlines

    # Set border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    amazon_orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    for region in regions:
        eks_client = boto3.client('eks', region_name=region)

        cluster_names = eks_client.list_clusters()['clusters']

        for cluster_name in cluster_names:
            cluster_data = {}
            cluster_data['cluster_name'] = cluster_name
            cluster_data['region'] = region

            response = eks_client.describe_cluster(name=cluster_name)
            cluster_data['version'] = response['cluster']['version']
            cluster_data['cluster_status'] = response['cluster']['status']
            cluster_data['vpc'] = response['cluster']['resourcesVpcConfig']['vpcId']

            nodegroups = []
            response = eks_client.list_nodegroups(clusterName=cluster_name)
            for nodegroup in response['nodegroups']:
                nodegroup_info = eks_client.describe_nodegroup(
                    clusterName=cluster_name,
                    nodegroupName=nodegroup
                )

                # Check if the instance type starts with 'g' or 'p' to determine if it's a GPU node
                instance_types = nodegroup_info['nodegroup']['instanceTypes']
                is_gpu_node = any(instance_type.startswith(('g', 'p')) for instance_type in instance_types)

                nodegroup_data = {
                    'name': nodegroup,
                    'node_instance_type': instance_types,
                    'ami_type': nodegroup_info['nodegroup']['amiType'],
                    'is_gpu_node': is_gpu_node,
                    'releaseVersion': nodegroup_info['nodegroup']['releaseVersion']
                }
                nodegroups.append(nodegroup_data)
            cluster_data['nodegroups'] = nodegroups

            addons = []

            response = eks_client.list_addons(clusterName=cluster_name)
            for addon in response['addons']:
                addon_info = eks_client.describe_addon(
                    clusterName=cluster_name,
                    addonName=addon
                )

                service_account_role_arn = addon_info['addon'].get('serviceAccountRoleArn', '')
                if service_account_role_arn:
                    service_account = service_account_role_arn.split('/')[-1]
                else:
                    service_account = ''

                addon_data = {
                    'name': addon,
                    'version': addon_info['addon']['addonVersion'],
                    'status': addon_info['addon']['status'],
                    'service_account': service_account,
                    'addon_pods': ''
                }

                addons.append(addon_data)

            deployments = os.environ.get('ADDON_CONTROLLER').split(',')
            namespace = os.environ.get('NAMESPACE')

            kubeconfig = get_cluster_conf(eks_client, cluster_name, region)

            config.load_kube_config_from_dict(config_dict=kubeconfig)
            v1_api = client.CoreV1Api()
            apps_api = client.AppsV1Api()

            for deployment in deployments:
                try:
                    deployment_info = apps_api.read_namespaced_deployment(deployment, namespace)

                    # Get the label selector for the Deployment
                    deployment_selector = deployment_info.spec.selector.match_labels

                    # Build the label selector string
                    label_selector_str = ','.join([f'{k}={v}' for k, v in deployment_selector.items()])

                    # Use the label selector to get the list of Pods associated with the Deployment
                    pod_list = v1_api.list_namespaced_pod(namespace, label_selector=label_selector_str)

                    # Extract the Pod names from the Pod list
                    deployment_pods = [pod.metadata.name for pod in pod_list.items]

                    # Get the available replicas and total replicas for the Deployment
                    available_replicas = deployment_info.status.available_replicas
                    replicas = deployment_info.status.replicas

                    addon_data = {
                        'name': deployment,
                        'version': '',
                        'status': f'Available: {available_replicas}/{replicas}',
                        'service_account': deployment_info.spec.template.spec.service_account_name,
                        'addon_pods': deployment_pods
                    }
                    addons.append(addon_data)
                except ApiException as e:
                    if e.status == 404:
                        print(f'Deployment {deployment} not found in cluster {cluster_name}')
                    else:
                        print(f"Error accessing deployment {deployment} in cluster {cluster_name}: {e}")
                except Exception as e:
                    print(f"Error accessing deployment {deployment} in cluster {cluster_name}: {e}")

            cluster_data['addons'] = addons

            # Write cluster information table
            headers_cluster = ['Cluster Name', 'Version', 'VPC', 'Region', 'Status']
            worksheet.append(headers_cluster)
            for cell in worksheet[worksheet._current_row]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                cell.fill = amazon_orange_fill  # Set header fill color
                cell.border = thin_border  # Set cell border

            row_cluster = [
                cluster_data['cluster_name'],
                cluster_data.get('version', ''),
                cluster_data.get('vpc', ''),
                cluster_data.get('region', ''),
                cluster_data.get('cluster_status', '')
            ]
            worksheet.append(row_cluster)
            for cell in worksheet[worksheet._current_row]:
                cell.alignment = Alignment(horizontal='left')  # Set row content left-aligned
                cell.border = thin_border  # Set cell border

            # Autofit column width
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            worksheet.append([])

            # Write NodeGroup table
            headers_nodegroups = ['NodeGroup Name', 'Version', 'Node Instance Type', 'AMI Type', 'GPU Node']
            worksheet.append(headers_nodegroups)
            for cell in worksheet[worksheet._current_row]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                cell.border = thin_border  # Set cell border

            for nodegroup in cluster_data.get('nodegroups', []):
                row_nodegroups = [
                    nodegroup['name'],
                    nodegroup['releaseVersion'],
                    ', '.join(nodegroup['node_instance_type']),
                    nodegroup['ami_type'],
                    nodegroup['is_gpu_node']
                ]
                worksheet.append(row_nodegroups)
                for cell in worksheet[worksheet._current_row]:
                    cell.alignment = Alignment(horizontal='left')  # Set row content left-aligned
                    cell.border = thin_border  # Set cell border

            # Autofit column width
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            worksheet.append([])

            # Write AddOn table
            headers_addons = ['AddOn Name', 'Version', 'Status', 'Service Account', 'AddOn Pods']
            worksheet.append(headers_addons)
            for cell in worksheet[worksheet._current_row]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                cell.border = thin_border  # Set cell border

            for addon in cluster_data.get('addons', []):
                row_addons = [
                    addon['name'],
                    addon['version'],
                    addon['status'],
                    addon['service_account'],
                    '\n'.join(addon['addon_pods'])
                ]
                worksheet.append(row_addons)
                for cell in worksheet[worksheet._current_row]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # Set row content left-aligned and wrap text
                    cell.border = thin_border  # Set cell border

            # Autofit column width, considering newline characters
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                newline_count = max(str(cell.value).count('\n') for cell in col)
                adjusted_width = (max_length) * 1.05 / (newline_count + 1)
                worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

            worksheet.append([])

    # Save the Excel file to memory
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Upload the Excel file to the S3 Bucket
    original_filename = 'eks-cluster-info/cluster_info_global.xlsx'
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    object_key = f"{original_filename.split('.')[0]}_{timestamp}.{original_filename.split('.')[-1]}"
    s3_client = boto3.client('s3')
    try:
        s3_client.upload_fileobj(output, bucket_name, object_key)
        print(f'Cluster information table saved to S3 bucket: {bucket_name}/{object_key}')
    except Exception as e:
        print(f'Error uploading file to S3: {e}')

    return {
        'statusCode': 200,
        'body': json.dumps('Lambda function executed successfully!')
    }